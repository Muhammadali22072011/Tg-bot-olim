import os
import sqlite3
import logging
import threading
import time
from datetime import datetime
from typing import List, Dict, Any, Optional

import telebot
from telebot import types
from openpyxl import Workbook
from openpyxl.styles import Font

# --- Configuration - Replace with your actual values ---

BOT_TOKEN = "8231571160:AAHgw1Dqrb4oAYN2euwq8OO20iC0tpmoBcI"
ADMIN_IDS = []  # Replace with actual admin Telegram IDs
SUPER_ADMIN_ID = 2050639074  # Replace with actual super admin ID

# --- Initialize bot ---

bot = telebot.TeleBot(BOT_TOKEN)

# --- Set up logging ---

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# --- Database setup ---

def init_db():
    """Initializes the SQLite database and tables."""
    conn = sqlite3.connect('olympiad_bot.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER UNIQUE NOT NULL,
            first_name TEXT,
            last_name TEXT,
            grade TEXT,
            phone TEXT,
            email TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS olympiads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            date TEXT,
            price REAL NOT NULL,
            banner_file_id TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS registrations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            olympiad_id INTEGER NOT NULL,
            payment_status TEXT NOT NULL,
            receipt_photo_id TEXT,
            admin_notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (olympiad_id) REFERENCES olympiads (id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER UNIQUE NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

init_db()

# --- User states for registration process ---

user_states = {}

# --- Helper functions for database operations ---

def get_db_connection():
    """Establishes and returns a database connection."""
    return sqlite3.connect('olympiad_bot.db')

def is_admin(telegram_id):
    """Checks if a user is an admin."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT 1 FROM admins WHERE telegram_id = ?', (telegram_id,))
    result = cursor.fetchone()
    conn.close()
    return result is not None

def is_super_admin(telegram_id):
    """Checks if a user is the super admin."""
    return telegram_id == SUPER_ADMIN_ID

def get_olympiads():
    """Fetches all olympiads from the database."""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM olympiads ORDER BY created_at DESC')
    olympiads = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return olympiads

def get_olympiad(olympiad_id):
    """Fetches a single olympiad by its ID."""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM olympiads WHERE id = ?', (olympiad_id,))
    result = cursor.fetchone()
    conn.close()
    return dict(result) if result else None

def get_user(telegram_id):
    """Fetches a user by their Telegram ID."""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE telegram_id = ?', (telegram_id,))
    result = cursor.fetchone()
    conn.close()
    return dict(result) if result else None

def get_user_by_id(user_id):
    """Fetches a user by their database ID."""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
    result = cursor.fetchone()
    conn.close()
    return dict(result) if result else None

def get_registration(user_id, olympiad_id):
    """Fetches a registration record for a specific user and olympiad."""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('''
        SELECT r.*, u.first_name, u.last_name, o.name as olympiad_name
        FROM registrations r
        JOIN users u ON r.user_id = u.id
        JOIN olympiads o ON r.olympiad_id = o.id
        WHERE u.id = ? AND o.id = ?
    ''', (user_id, olympiad_id))
    result = cursor.fetchone()
    conn.close()
    return dict(result) if result else None

def get_olympiad_participants(olympiad_id):
    """Fetches all participants for a specific olympiad."""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('''
        SELECT u.*, r.payment_status, r.created_at as registration_date
        FROM registrations r
        JOIN users u ON r.user_id = u.id
        WHERE r.olympiad_id = ?
    ''', (olympiad_id,))
    participants = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return participants

def get_pending_payments():
    """Fetches all payments that are waiting for review."""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('''
        SELECT r.*, u.first_name, u.last_name, u.telegram_id, o.name as olympiad_name
        FROM registrations r
        JOIN users u ON r.user_id = u.id
        JOIN olympiads o ON r.olympiad_id = o.id
        WHERE r.payment_status IN ('waiting_receipt', 'pending')
        ORDER BY r.created_at DESC
    ''')
    payments = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return payments

def update_payment_status(registration_id, status, admin_notes=None):
    """Updates the payment status of a registration."""
    conn = get_db_connection()
    cursor = conn.cursor()
    if admin_notes:
        cursor.execute('''
            UPDATE registrations
            SET payment_status = ?, admin_notes = ?
            WHERE id = ?
        ''', (status, admin_notes, registration_id))
    else:
        cursor.execute('''
            UPDATE registrations
            SET payment_status = ?
            WHERE id = ?
        ''', (status, registration_id))
    conn.commit()
    conn.close()

def get_all_admins():
    """Fetches all admin IDs from the database."""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM admins')
    admins = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return admins

def clear_user_state(user_id):
    """Clears the state for a specific user."""
    if user_id in user_states:
        del user_states[user_id]

# --- Excel export function ---

def export_to_excel(olympiad_id, olympiad_name):
    """Exports participants data for a specific olympiad to an Excel file."""
    try:
        participants = get_olympiad_participants(olympiad_id)

        wb = Workbook()
        ws = wb.active
        
        # Limit sheet name to 31 characters
        sheet_name = f"Participants_{olympiad_name}"[:30]
        # Replace invalid characters
        invalid_chars = '/\*?[]'
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '')
        ws.title = sheet_name
        
        # Add headers
        headers = ["ID", "Ism", "Familiya", "Sinf", "Telefon", "Email", "To'lov holati", "Ro'yxatdan o'tgan sana"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        # Add data
        for row, participant in enumerate(participants, 2):
            ws.cell(row=row, column=1, value=participant['id'])
            ws.cell(row=row, column=2, value=participant['first_name'])
            ws.cell(row=row, column=3, value=participant['last_name'])
            ws.cell(row=row, column=4, value=participant['grade'])
            ws.cell(row=row, column=5, value=participant['phone'])
            ws.cell(row=row, column=6, value=participant['email'])
            ws.cell(row=row, column=7, value=participant['payment_status'])
            ws.cell(row=row, column=8, value=participant['registration_date'])
        
        # Save the file
        filename = f"participants_{olympiad_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # Sanitize filename
        filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_', '.')).rstrip()
        
        wb.save(filename)
        return filename

    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        return None

# --- Participant handlers ---

@bot.message_handler(commands=['start'])
def cmd_start(message):
    """Handles the /start command, showing available olympiads."""
    user_id = message.from_user.id
    clear_user_state(user_id)
    
    olympiads = get_olympiads()
    
    keyboard = types.InlineKeyboardMarkup()
    for o in olympiads:
        keyboard.add(types.InlineKeyboardButton(f"🏆 {o['name']}", callback_data=f"olympiad_{o['id']}"))
    
    bot.send_message(
        message.chat.id,
        "Assalomu alaykum! 👋 Olimpiadaga xush kelibsiz. Iltimos, olimpiadani tanlang:",
        reply_markup=keyboard
    )
    
@bot.callback_query_handler(func=lambda call: call.data.startswith('olympiad_'))
def process_olympiad_choice(call):
    """Shows details of the selected olympiad."""
    try:
        olympiad_id = int(call.data.split("_")[1])
        olympiad = get_olympiad(olympiad_id)
        user_id = call.from_user.id
        
        user_states[user_id] = {'state': 'choosing_olympiad', 'olympiad_id': olympiad_id}
        
        if olympiad:
            message_text = (
                f"<b>🏆 {olympiad['name']}</b>\n\n"
                f"<i>Tavsif:</i> {olympiad['description']}\n"
                f"<i>Sana:</i> {olympiad['date']}\n"
                f"<i>Narxi:</i> {olympiad['price']:,} so'm"
            )
            
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(types.InlineKeyboardButton("✅ Ro'yxatdan o'tish", callback_data="register"))
            
            # --- FIX: Use .get() to safely access 'banner_file_id' ---
            banner_file_id = olympiad.get('banner_file_id')
            if banner_file_id:
                bot.send_photo(
                    call.message.chat.id,
                    banner_file_id,
                    caption=message_text,
                    reply_markup=keyboard,
                    parse_mode='HTML'
                )
            else:
                bot.send_message(
                    call.message.chat.id,
                    message_text,
                    reply_markup=keyboard,
                    parse_mode='HTML'
                )
        
        bot.answer_callback_query(call.id)
        
    except Exception as e:
        logger.error(f"Error in process_olympiad_choice: {e}")
        bot.send_message(call.message.chat.id, "Kechirasiz, tanlangan olimpiada ma'lumotlarini yuklashda xatolik yuz berdi.")
        bot.answer_callback_query(call.id, "Xatolik yuz berdi.", show_alert=True)

@bot.callback_query_handler(func=lambda call: call.data == 'register')
def start_registration(call):
    """Starts the registration process."""
    user_id = call.from_user.id
    user_state = user_states.get(user_id, {})
    olympiad_id = user_state.get('olympiad_id')
    
    if not olympiad_id:
        bot.send_message(call.message.chat.id, "Iltimos, avval olimpiadani tanlang.")
        return
        
    user = get_user(user_id)
    
    if user:
        # User already exists, check for registration
        registration = get_registration(user['id'], olympiad_id)
        if registration:
            status = registration['payment_status']
            if status == 'approved':
                bot.send_message(call.message.chat.id, "Siz bu olimpiadaga allaqachon ro'yxatdan o'tgansiz va to'lovingiz tasdiqlangan. ✅")
            elif status == 'pending' or status == 'waiting_receipt':
                bot.send_message(call.message.chat.id, "Sizning ro'yxatdan o'tishingiz ko'rib chiqilmoqda. Iltimos, kuting... ⏳")
            else:
                # Registration exists but was rejected
                olympiad = get_olympiad(olympiad_id)
                user_states[user_id]['state'] = 'sending_receipt'
                message_text = (
                    f"Siz avvalroq ro'yxatdan o'tgan edingiz, lekin to'lovingiz rad etilgan. Iltimos, to'lovni qayta amalga oshiring. 😥\n"
                    f"Narxi: {olympiad['price']:,} so'm\n\n"
                    f"Iltimos, to'lov chekini (screenshot) yuboring. To'lov quyidagi kartaga o'tkaziladi:\n"
                    f"💳 <code>8600 1234 5678 9010</code>\n"
                    f"👤 OLIMPIADA MARKAZI"
                )
                bot.send_message(call.message.chat.id, message_text, parse_mode='HTML')
        else:
            # User exists but not registered for this olympiad
            user_states[user_id]['state'] = 'confirming_details'
            bot.send_message(call.message.chat.id, "Siz allaqachon ro'yxatdan o'tgansiz. Davom etish uchun /start buyrug'ini bosing")
            
    else:
        # New user
        user_states[user_id]['state'] = 'entering_first_name'
        bot.send_message(call.message.chat.id, "Ro'yxatdan o'tishni boshlaymiz. Iltimos, <b>Ismingizni</b> kiriting:", parse_mode='HTML')

    bot.answer_callback_query(call.id)

@bot.message_handler(func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'entering_first_name')
def process_first_name(message):
    user_id = message.from_user.id
    user_state = user_states.get(user_id, {})
    user_state['first_name'] = message.text
    user_state['state'] = 'entering_last_name'
    user_states[user_id] = user_state
    
    bot.send_message(message.chat.id, "<b>Familiyangizni</b> kiriting:", parse_mode='HTML')

@bot.message_handler(func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'entering_last_name')
def process_last_name(message):
    user_id = message.from_user.id
    user_state = user_states.get(user_id, {})
    user_state['last_name'] = message.text
    user_state['state'] = 'entering_grade'
    user_states[user_id] = user_state
    
    bot.send_message(message.chat.id, "<b>Sinfingizni</b> tanlang (5-11):", parse_mode='HTML')

@bot.message_handler(func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'entering_grade')
def process_grade(message):
    user_id = message.from_user.id
    user_state = user_states.get(user_id, {})
    try:
        grade = int(message.text)
        if 5 <= grade <= 11:
            user_state['grade'] = grade
            user_state['state'] = 'entering_phone'
            user_states[user_id] = user_state
            
            keyboard = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
            keyboard.add(types.KeyboardButton("📞 Telefon raqamni jo'natish", request_contact=True))
            bot.send_message(message.chat.id, "Endi <b>telefon raqamingizni</b> jo'nating:", reply_markup=keyboard, parse_mode='HTML')
        else:
            bot.send_message(message.chat.id, "Iltimos, 5 dan 11 gacha bo'lgan raqam kiriting: 🔢")
    except ValueError:
        bot.send_message(message.chat.id, "Iltimos, 5 dan 11 gacha bo'lgan raqam kiriting: 🔢")

@bot.message_handler(content_types=['contact'], func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'entering_phone')
def process_phone_contact(message):
    user_id = message.from_user.id
    user_state = user_states.get(user_id, {})
    user_state['phone'] = message.contact.phone_number
    user_state['state'] = 'entering_email'
    user_states[user_id] = user_state
    
    bot.send_message(
        message.chat.id,
        "<b>Email manzilingiz</b> yoki <b>Telegram foydalanuvchi nomingizni</b> kiriting:",
        reply_markup=types.ReplyKeyboardRemove(),
        parse_mode='HTML'
    )

@bot.message_handler(func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'entering_phone')
def process_phone_text(message):
    user_id = message.from_user.id
    user_state = user_states.get(user_id, {})
    user_state['phone'] = message.text
    user_state['state'] = 'entering_email'
    user_states[user_id] = user_state
    
    bot.send_message(
        message.chat.id,
        "<b>Email manzilingiz</b> yoki <b>Telegram foydalanuvchi nomingizni</b> kiriting:",
        reply_markup=types.ReplyKeyboardRemove(),
        parse_mode='HTML'
    )

@bot.message_handler(func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'entering_email')
def process_email(message):
    user_id = message.from_user.id
    user_state = user_states.get(user_id, {})
    user_state['email'] = message.text
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check if user already exists
    user = get_user(user_id)
    if not user:
        cursor.execute('''
            INSERT INTO users (telegram_id, first_name, last_name, grade, phone, email)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            user_id,
            user_state['first_name'],
            user_state['last_name'],
            user_state['grade'],
            user_state['phone'],
            user_state['email']
        ))
        conn.commit()
        user_db = get_user(user_id)
        user_id_db = user_db['id']
    else:
        # Update existing user info
        cursor.execute('''
            UPDATE users SET first_name=?, last_name=?, grade=?, phone=?, email=? WHERE telegram_id=?
        ''', (
            user_state['first_name'],
            user_state['last_name'],
            user_state['grade'],
            user_state['phone'],
            user_state['email'],
            user_id
        ))
        conn.commit()
        user_id_db = user['id']

    olympiad_id = user_state.get('olympiad_id')
    olympiad = get_olympiad(olympiad_id)

    # Insert registration record
    cursor.execute('''
        INSERT INTO registrations (user_id, olympiad_id, payment_status)
        VALUES (?, ?, ?)
    ''', (user_id_db, olympiad_id, 'pending'))
    conn.commit()
    conn.close()
    
    user_state['state'] = 'sending_receipt'
    user_state['registration_id'] = cursor.lastrowid
    user_states[user_id] = user_state

    message_text = (
        f"✅ Ro'yxatdan o'tish yakunlandi. Endi to'lov qilishingiz kerak. \n"
        f"<b>Narxi:</b> {olympiad['price']:,} so'm\n\n"
        f"Iltimos, to'lov chekini (screenshot) yuboring. To'lov quyidagi kartaga o'tkaziladi:\n"
        f"💳 <code>8600 1234 5678 9010</code>\n"
        f"👤 OLIMPIADA MARKAZI"
    )
    bot.send_message(message.chat.id, message_text, parse_mode='HTML')

@bot.message_handler(content_types=['photo'], func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'sending_receipt')
def process_receipt_photo(message):
    user_id = message.from_user.id
    user_state = user_states.get(user_id, {})
    registration_id = user_state.get('registration_id')

    if not registration_id:
        bot.send_message(message.chat.id, "Kechirasiz, siz ro'yxatdan o'tmagansiz. Iltimos, /start buyrug'ini bosing.")
        return

    # Update payment status and save receipt photo ID
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE registrations
        SET payment_status = 'waiting_receipt', receipt_photo_id = ?
        WHERE id = ?
    ''', (message.photo[-1].file_id, registration_id))
    conn.commit()
    conn.close()
    
    bot.send_message(message.chat.id, "✅ To'lov cheki qabul qilindi. Administrator tomonidan tekshirilgach, sizga xabar beramiz. Tasdiqlanishini kuting. ⏳")
    
    registration = get_registration_by_id(registration_id)
    if registration:
        notify_admins_of_new_receipt(registration)
        
    clear_user_state(user_id)

def notify_admins_of_new_receipt(registration):
    """Sends a notification to all admins about a new payment receipt."""
    admins = get_all_admins()
    user = get_user_by_id(registration['user_id'])
    olympiad = get_olympiad(registration['olympiad_id'])
    
    message_text = (
        "🆕 **Yangi to'lov cheki!**\n\n"
        f"👤 Ismi: {user['first_name']} {user['last_name']}\n"
        f"📞 Telefon: {user['phone']}\n"
        f"📚 Olimpiada: {olympiad['name']}\n"
        f"💳 Summa: {olympiad['price']:,} so'm\n\n"
        f"Tekshirish uchun /admin buyrug'idan foydalaning. 🔍"
    )
    
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(types.InlineKeyboardButton("➡️ Tekshirishga o'tish", callback_data="admin_check_payments"))
    
    for admin in admins:
        try:
            bot.send_photo(
                admin['telegram_id'],
                registration['receipt_photo_id'],
                caption=message_text,
                reply_markup=keyboard,
                parse_mode='Markdown'
            )
        except Exception as e:
            logger.error(f"Failed to send payment notification to admin {admin['telegram_id']}: {e}")

# --- Admin handlers ---

@bot.message_handler(commands=['admin'])
def cmd_admin(message):
    """Shows the admin panel."""
    if not is_admin(message.from_user.id):
        bot.send_message(message.chat.id, "Sizda admin huquqi yo'q. 🚫")
        return
        
    clear_user_state(message.from_user.id)
    
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(types.InlineKeyboardButton("📄 To'lovlarni tekshirish", callback_data="admin_check_payments"))
    keyboard.add(types.InlineKeyboardButton("🏆 Olimpiadalarni boshqarish", callback_data="admin_manage_olympiads"))
    keyboard.add(types.InlineKeyboardButton("📝 Xabar yuborish", callback_data="admin_broadcast"))
    
    if is_super_admin(message.from_user.id):
        keyboard.add(types.InlineKeyboardButton("➕ Admin qo'shish", callback_data="admin_add_admin"))
        keyboard.add(types.InlineKeyboardButton("📊 Barcha Excel fayllarni eksport qilish", callback_data="admin_export_excel"))
        
    bot.send_message(message.chat.id, "Admin panelga xush kelibsiz. Quyidagi amallardan birini tanlang:", reply_markup=keyboard)

@bot.callback_query_handler(func=lambda call: call.data == 'admin_check_payments')
def admin_check_payments(call):
    """Starts the payment review process."""
    if not is_admin(call.from_user.id):
        bot.send_message(call.message.chat.id, "Sizda admin huquqi yo'q. 🚫")
        return
        
    pending_payments = get_pending_payments()
    
    if not pending_payments:
        bot.send_message(call.message.chat.id, "Hozircha tekshirish kerak bo'lgan to'lovlar mavjud emas. ✅")
        bot.answer_callback_query(call.id, "Tekshiriladigan to'lovlar yo'q.", show_alert=True)
        return
    
    user_id = call.from_user.id
    user_states[user_id] = {
        'state': 'reviewing_payments',
        'payments': pending_payments,
        'current_index': 0
    }
    
    show_payment_for_review(
        call.message.chat.id,
        pending_payments[0],
        0,
        len(pending_payments)
    )
    bot.answer_callback_query(call.id)
    
def get_registration_by_id(registration_id):
    """Fetches a registration by its ID."""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM registrations WHERE id = ?', (registration_id,))
    result = cursor.fetchone()
    conn.close()
    return dict(result) if result else None

def show_payment_for_review(chat_id, payment, index, total):
    """Displays a payment for admin review."""
    user = get_user_by_id(payment['user_id'])
    olympiad = get_olympiad(payment['olympiad_id'])
    
    message_text = (
        f"<b>To'lov #{index + 1} / {total}</b>\n\n"
        f"👤 Ismi: {user['first_name']} {user['last_name']}\n"
        f"📞 Telefon: {user['phone']}\n"
        f"📧 Email: {user['email']}\n"
        f"📚 Sinf: {user['grade']}\n\n"
        f"🏆 Olimpiada: {olympiad['name']}\n"
        f"💳 Summa: {olympiad['price']:,} so'm\n"
        f"📆 Ro'yxatdan o'tgan: {payment['created_at']}\n"
        f"🔎 Holat: {payment['payment_status']}"
    )
    
    keyboard = types.InlineKeyboardMarkup()
    keyboard.row(
        types.InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve_{payment['id']}_{index}"),
        types.InlineKeyboardButton("❌ Rad etish", callback_data=f"reject_{payment['id']}_{index}")
    )
    
    if total > 1:
        keyboard.row(
            types.InlineKeyboardButton("⏮ Oldingi", callback_data=f"prev_{index}"),
            types.InlineKeyboardButton("⏭ Keyingi", callback_data=f"next_{index}")
        )
    
    try:
        if payment['receipt_photo_id']:
            bot.send_photo(chat_id, payment['receipt_photo_id'], caption=message_text, reply_markup=keyboard, parse_mode='HTML')
        else:
            bot.send_message(chat_id, message_text, reply_markup=keyboard, parse_mode='HTML')
    except Exception as e:
        logger.error(f"Error showing payment: {e}")
        bot.send_message(chat_id, "Xatolik yuz berdi. Qayta urinib ko'ring.")
        
@bot.callback_query_handler(func=lambda call: call.data.startswith('approve_') or call.data.startswith('reject_'))
def handle_payment_decision(call):
    """Handles the approval or rejection of a payment."""
    if not is_admin(call.from_user.id):
        bot.send_message(call.message.chat.id, "Sizda admin huquqi yo'q. 🚫")
        return
    
    parts = call.data.split('_')
    action = parts[0]
    registration_id = int(parts[1])
    current_index = int(parts[2])

    registration = get_registration_by_id(registration_id)
    if not registration:
        bot.send_message(call.message.chat.id, "Xatolik: to'lov topilmadi. 🤷‍♂️")
        bot.answer_callback_query(call.id, "Xatolik: to'lov topilmadi.", show_alert=True)
        return

    user_id = registration['user_id']
    user_info = get_user_by_id(user_id)
    olympiad_info = get_olympiad(registration['olympiad_id'])

    if action == 'approve':
        update_payment_status(registration_id, 'approved', f"Tasdiqlangan. Admin: {call.from_user.id}")
        bot.send_message(call.message.chat.id, "✅ To'lov tasdiqlandi va foydalanuvchi xabar oldi.")
        try:
            bot.send_message(
                user_info['telegram_id'],
                f"✅ Sizning to'lovingiz tasdiqlandi!\n\n"
                f"Siz <b>{olympiad_info['name']}</b> olimpiadasiga muvaffaqiyatli ro'yxatdan o'tdingiz. Ishtirokingiz uchun tashakkur! 😊\n\n"
                f"Olimpiada haqida batafsil ma'lumot keyinroq yuboriladi.",
                parse_mode='HTML'
            )
        except Exception as e:
            logger.error(f"Failed to notify user {user_info['telegram_id']}: {e}")
            bot.send_message(call.message.chat.id, f"Foydalanuvchiga xabar yuborishda xatolik: {e}")
        
        # Move to the next payment if available
        pending_payments = get_pending_payments()
        if pending_payments:
            if current_index < len(pending_payments):
                show_payment_for_review(call.message.chat.id, pending_payments[current_index], current_index, len(pending_payments))
            else:
                bot.send_message(call.message.chat.id, "Barcha to'lovlar tekshirildi. ✅")
        else:
            bot.send_message(call.message.chat.id, "Hozircha tekshirish kerak bo'lgan to'lovlar mavjud emas. ✅")

    elif action == 'reject':
        user_states[call.from_user.id] = {
            'state': 'rejecting_payment',
            'registration_id': registration_id,
            'current_index': current_index
        }
        msg = bot.send_message(call.message.chat.id, "❌ Iltimos, to'lovni rad etish sababini kiriting:")
        bot.register_next_step_handler(msg, process_rejection_reason, registration_id, current_index)
        
    bot.answer_callback_query(call.id)

def process_rejection_reason(message, registration_id, current_index):
    """Processes the reason for payment rejection."""
    user_id = message.from_user.id
    user_state = user_states.get(user_id, {})
    if user_state.get('state') != 'rejecting_payment':
        return
        
    reason = message.text
    update_payment_status(registration_id, 'rejected', f"Rad etilgan: {reason}. Admin: {user_id}")
    
    registration = get_registration_by_id(registration_id)
    if registration:
        user_info = get_user_by_id(registration['user_id'])
        
        try:
            bot.send_message(
                user_info['telegram_id'],
                f"❌ Sizning to'lovingiz rad etildi. \n\n"
                f"<b>Sabab:</b> {reason}\n\n"
                f"Iltimos, to'g'ri to'lov cheki yuboring yoki biz bilan bog'laning. 🤔",
                parse_mode='HTML'
            )
        except Exception as e:
            logger.error(f"Failed to notify user {user_info['telegram_id']}: {e}")
            
    bot.send_message(message.chat.id, "❌ To'lov rad etildi va foydalanuvchi xabar oldi.")
    
    clear_user_state(user_id)
    
    # Show the next payment
    pending_payments = get_pending_payments()
    if pending_payments:
        if current_index < len(pending_payments):
            show_payment_for_review(message.chat.id, pending_payments[current_index], current_index, len(pending_payments))
        else:
            bot.send_message(message.chat.id, "Barcha to'lovlar tekshirildi. ✅")
    else:
        bot.send_message(message.chat.id, "Hozircha tekshirish kerak bo'lgan to'lovlar mavjud emas. ✅")

@bot.callback_query_handler(func=lambda call: call.data.startswith('prev_') or call.data.startswith('next_'))
def navigate_payments(call):
    """Handles navigation between pending payments."""
    if not is_admin(call.from_user.id):
        bot.send_message(call.message.chat.id, "Sizda admin huquqi yo'q. 🚫")
        return
    
    action = call.data.split('_')[0]
    current_index = int(call.data.split('_')[1])
    
    pending_payments = get_pending_payments()
    
    if not pending_payments:
        bot.send_message(call.message.chat.id, "Hozircha tekshirish kerak bo'lgan to'lovlar mavjud emas. ✅")
        bot.answer_callback_query(call.id, "To'lovlar topilmadi.", show_alert=True)
        return
    
    if action == 'prev':
        new_index = (current_index - 1 + len(pending_payments)) % len(pending_payments)
    else:  # 'next'
        new_index = (current_index + 1) % len(pending_payments)
    
    # Delete old message to avoid spam
    try:
        bot.delete_message(call.message.chat.id, call.message.message_id)
    except Exception as e:
        logger.warning(f"Failed to delete message: {e}")
    
    # Show the new payment
    show_payment_for_review(call.message.chat.id, pending_payments[new_index], new_index, len(pending_payments))
    
    bot.answer_callback_query(call.id)
    
@bot.callback_query_handler(func=lambda call: call.data == 'admin_manage_olympiads')
def admin_manage_olympiads(call):
    """Shows options for managing olympiads."""
    if not is_admin(call.from_user.id):
        bot.send_message(call.message.chat.id, "Sizda admin huquqi yo'q. 🚫")
        return
        
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(types.InlineKeyboardButton("➕ Yangi olimpiada qo'shish", callback_data="admin_add_olympiad"))
    keyboard.add(types.InlineKeyboardButton("🏆 Olimpiadalarni ko'rish", callback_data="admin_view_olympiads"))
    
    bot.send_message(call.message.chat.id, "Olimpiadalarni boshqarish amalini tanlang:", reply_markup=keyboard)
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data == 'admin_add_olympiad')
def admin_add_olympiad(call):
    """Starts the process of adding a new olympiad."""
    if not is_admin(call.from_user.id):
        bot.send_message(call.message.chat.id, "Sizda admin huquqi yo'q. 🚫")
        return
        
    clear_user_state(call.from_user.id)
    
    user_states[call.from_user.id] = {'state': 'adding_olympiad'}
    msg = bot.send_message(call.message.chat.id, "Yangi olimpiada <b>nomini</b> kiriting:", parse_mode='HTML')
    bot.register_next_step_handler(msg, process_olympiad_name)
    bot.answer_callback_query(call.id)

def process_olympiad_name(message):
    user_states[message.from_user.id]['name'] = message.text
    user_states[message.from_user.id]['state'] = 'adding_olympiad_description'
    msg = bot.send_message(message.chat.id, "Olimpiada haqida <b>tavsif</b> kiriting:", parse_mode='HTML')
    bot.register_next_step_handler(msg, process_olympiad_description)

def process_olympiad_description(message):
    user_states[message.from_user.id]['description'] = message.text
    user_states[message.from_user.id]['state'] = 'adding_olympiad_date'
    msg = bot.send_message(message.chat.id, "Olimpiada <b>sanasini</b> kiriting (masalan: 15-sentyabr, soat 10:00):", parse_mode='HTML')
    bot.register_next_step_handler(msg, process_olympiad_date)

def process_olympiad_date(message):
    user_states[message.from_user.id]['date'] = message.text
    user_states[message.from_user.id]['state'] = 'adding_olympiad_price'
    msg = bot.send_message(message.chat.id, "Olimpiada <b>narxini</b> kiriting (so'mda, faqat raqam):", parse_mode='HTML')
    bot.register_next_step_handler(msg, process_olympiad_price)

def process_olympiad_price(message):
    user_id = message.from_user.id
    user_state = user_states.get(user_id, {})
    try:
        price = float(message.text)
        user_state['price'] = price
        user_states[user_id]['state'] = 'adding_olympiad_banner'
        msg = bot.send_message(message.chat.id, "Olimpiada <b>banner rasmini</b> yuboring (file sifatida):", parse_mode='HTML')
        bot.register_next_step_handler(msg, process_olympiad_banner)
    except ValueError:
        msg = bot.send_message(message.chat.id, "Iltimos, faqat raqam kiriting. 🔢")
        bot.register_next_step_handler(msg, process_olympiad_price)

@bot.message_handler(content_types=['photo'], func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'adding_olympiad_banner')
def process_olympiad_banner(message):
    user_id = message.from_user.id
    user_state = user_states.get(user_id, {})
    
    olympiad_name = user_state.get('name')
    olympiad_description = user_state.get('description')
    olympiad_date = user_state.get('date')
    olympiad_price = user_state.get('price')
    banner_file_id = message.photo[-1].file_id

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO olympiads (name, description, date, price, banner_file_id)
        VALUES (?, ?, ?, ?, ?)
    ''', (olympiad_name, olympiad_description, olympiad_date, olympiad_price, banner_file_id))
    conn.commit()
    conn.close()
    
    bot.send_message(message.chat.id, "✅ Olimpiada muvaffaqiyatli qo'shildi!")
    clear_user_state(user_id)

@bot.message_handler(func=lambda message: user_states.get(message.from_user.id, {}).get('state') == 'adding_olympiad_banner' and message.content_type != 'photo')
def handle_invalid_banner(message):
    msg = bot.send_message(message.chat.id, "Noto'g'ri format. Iltimos, rasm faylini yuboring.")
    bot.register_next_step_handler(msg, process_olympiad_banner)

@bot.callback_query_handler(func=lambda call: call.data == 'admin_view_olympiads')
def admin_view_olympiads(call):
    """Displays a list of all olympiads for admins."""
    if not is_admin(call.from_user.id):
        bot.send_message(call.message.chat.id, "Sizda admin huquqi yo'q. 🚫")
        return
        
    olympiads = get_olympiads()
    
    if not olympiads:
        bot.send_message(call.message.chat.id, "Hozircha hech qanday olimpiada qo'shilmagan. 🤷‍♂️")
        return
        
    message_text = "<b>Olimpiadalar ro'yxati:</b>\n\n"
    keyboard = types.InlineKeyboardMarkup()
    for o in olympiads:
        participants_count = len(get_olympiad_participants(o['id']))
        message_text += f"🏆 {o['name']} (ID: {o['id']}) - {participants_count} ta ishtirokchi\n"
        keyboard.add(types.InlineKeyboardButton(f"📄 {o['name']} ishtirokchilari", callback_data=f"admin_participants_{o['id']}"))
    
    bot.send_message(call.message.chat.id, message_text, reply_markup=keyboard, parse_mode='HTML')
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data.startswith('admin_participants_'))
def admin_view_participants(call):
    """Displays participants for a specific olympiad and provides an export option."""
    if not is_admin(call.from_user.id):
        bot.send_message(call.message.chat.id, "Sizda admin huquqi yo'q. 🚫")
        return
        
    olympiad_id = int(call.data.split("_")[2])
    olympiad = get_olympiad(olympiad_id)
    participants = get_olympiad_participants(olympiad_id)
    
    if not participants:
        bot.send_message(call.message.chat.id, f"<b>{olympiad['name']}</b> olimpiadasi uchun hali ishtirokchilar yo'q. 🤷‍♀️", parse_mode='HTML')
        bot.answer_callback_query(call.id)
        return
        
    message_text = f"<b>{olympiad['name']}</b> olimpiadasi ishtirokchilari:\n\n"
    for p in participants:
        message_text += f"👤 {p['first_name']} {p['last_name']} ({p['grade']}-sinf)\n"
        message_text += f"📞 {p['phone']}\n"
        message_text += f"🔍 Holat: {p['payment_status']}\n\n"
        
    keyboard = types.InlineKeyboardMarkup()
    keyboard.add(types.InlineKeyboardButton("📊 Excelga eksport qilish", callback_data=f"admin_export_{olympiad_id}"))
    
    bot.send_message(call.message.chat.id, message_text, reply_markup=keyboard, parse_mode='HTML')
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data.startswith('admin_export_'))
def admin_export_excel(call):
    """Exports a single olympiad's participants to an Excel file."""
    if not is_admin(call.from_user.id):
        bot.send_message(call.message.chat.id, "Sizda admin huquqi yo'q. 🚫")
        return
        
    olympiad_id = int(call.data.split("_")[2])
    olympiad = get_olympiad(olympiad_id)
    
    try:
        file_path = export_to_excel(olympiad_id, olympiad['name'])
        if file_path and os.path.exists(file_path):
            with open(file_path, 'rb') as f:
                bot.send_document(call.message.chat.id, f, caption=f"✅ <b>{olympiad['name']}</b> olimpiadasi ishtirokchilari ro'yxati.", parse_mode='HTML')
            os.remove(file_path)
        else:
            bot.send_message(call.message.chat.id, "Excel faylini yaratishda xatolik yuz berdi. 😔")
            
    except Exception as e:
        logger.error(f"Error sending Excel file: {e}")
        bot.send_message(call.message.chat.id, "Excel faylini yuborishda xatolik yuz berdi. 😔")
        
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data == 'admin_export_excel')
def admin_export_all_excel(call):
    """Exports all olympiad participants to a single Excel file (if logic is implemented)."""
    if not is_super_admin(call.from_user.id):
        bot.send_message(call.message.chat.id, "Sizda super admin huquqi yo'q. 🚫")
        return
        
    bot.send_message(call.message.chat.id, "Bu funksiya hali to'liq ishga tushirilmagan. 🛠️")
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data == 'admin_add_admin')
def admin_add_admin(call):
    """Starts the process of adding a new admin."""
    if not is_super_admin(call.from_user.id):
        bot.send_message(call.message.chat.id, "Sizda super admin huquqi yo'q. 🚫")
        return
        
    clear_user_state(call.from_user.id)
    user_states[call.from_user.id] = {'state': 'adding_admin'}
    msg = bot.send_message(call.message.chat.id, "Yangi adminning <b>Telegram ID</b> sini kiriting:", parse_mode='HTML')
    bot.register_next_step_handler(msg, process_new_admin_id)
    bot.answer_callback_query(call.id)

def process_new_admin_id(message):
    user_id = message.from_user.id
    if user_states.get(user_id, {}).get('state') != 'adding_admin':
        return
    
    try:
        new_admin_id = int(message.text)
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT 1 FROM admins WHERE telegram_id = ?', (new_admin_id,))
        if cursor.fetchone():
            bot.send_message(message.chat.id, "Bu foydalanuvchi allaqachon admin hisoblanadi. 🤔")
        else:
            cursor.execute('INSERT INTO admins (telegram_id) VALUES (?)', (new_admin_id,))
            conn.commit()
            bot.send_message(message.chat.id, f"✅ Yangi admin (ID: {new_admin_id}) muvaffaqiyatli qo'shildi!")
            
        conn.close()
        clear_user_state(user_id)
    except ValueError:
        msg = bot.send_message(message.chat.id, "Iltimos, to'g'ri ID kiriting (faqat raqamlar). 🔢")
        bot.register_next_step_handler(msg, process_new_admin_id)

# --- Reminder system ---

def send_reminders():
    while True:
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Get upcoming olympiads (e.g., within 24 hours)
            cursor.execute('''
                SELECT * FROM olympiads
                WHERE date > CURRENT_TIMESTAMP AND strftime('%s', date) - strftime('%s', CURRENT_TIMESTAMP) < 86400
            ''')
            upcoming_olympiads = cursor.fetchall()
            
            for o in upcoming_olympiads:
                # Find users who have registered but not paid
                cursor.execute('''
                    SELECT u.telegram_id FROM registrations r
                    JOIN users u ON r.user_id = u.id
                    WHERE r.olympiad_id = ? AND r.payment_status IN ('pending', 'waiting_receipt')
                ''', (o[0],))
                
                users_to_remind = [row[0] for row in cursor.fetchall()]
                
                for user_telegram_id in users_to_remind:
                    try:
                        bot.send_message(
                            user_telegram_id,
                            f"🔔 Esingizda tuting! <b>{o[1]}</b> olimpiadasiga ro'yxatdan o'tgan edingiz. \n"
                            f"Iltimos, ishtirokingizni tasdiqlash uchun to'lovni yakunlang. 😊",
                            parse_mode='HTML'
                        )
                        time.sleep(1) # To avoid rate limits
                    except Exception as e:
                        logger.error(f"Failed to send reminder to {user_telegram_id}: {e}")
                        
            conn.close()
            time.sleep(3600 * 6) # Check every 6 hours
        except Exception as e:
            logger.error(f"Error in reminder thread: {e}")
            time.sleep(3600) # Wait 1 hour before retrying

# --- Broadcast handlers ---

@bot.callback_query_handler(func=lambda call: call.data == 'admin_broadcast')
def admin_broadcast(call):
    if not is_admin(call.from_user.id):
        bot.send_message(call.message.chat.id, "Sizda admin huquqi yo'q. 🚫")
        return
        
    clear_user_state(call.from_user.id)
    
    olympiads = get_olympiads()
    
    keyboard = types.InlineKeyboardMarkup()
    for o in olympiads:
        keyboard.add(types.InlineKeyboardButton(f"🏆 {o['name']}", callback_data=f"broadcast_{o['id']}"))
    
    keyboard.add(types.InlineKeyboardButton("🌐 Barcha foydalanuvchilarga", callback_data="broadcast_all"))
    
    bot.send_message(call.message.chat.id, "Qaysi olimpiada ishtirokchilariga xabar yubormoqchisiz? 📢", reply_markup=keyboard)
    bot.answer_callback_query(call.id)

@bot.callback_query_handler(func=lambda call: call.data.startswith('broadcast_'))
def admin_broadcast_olympiad(call):
    target = call.data.split('_')[1]
    
    user_states[call.from_user.id] = {
        'state': 'broadcasting',
        'target_id': target
    }
    
    msg = bot.send_message(call.message.chat.id, "Yubormoqchi bo'lgan xabaringizni kiriting: 📝")
    bot.register_next_step_handler(msg, admin_process_broadcast)
    bot.answer_callback_query(call.id)

def admin_process_broadcast(message):
    user_id = message.from_user.id
    user_state = user_states.get(user_id, {})
    target_id = user_state.get('target_id')
    
    if user_state.get('state') != 'broadcasting':
        return
        
    text_to_send = message.text
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if target_id == 'all':
        cursor.execute('SELECT telegram_id FROM users')
    else:
        try:
            olympiad_id = int(target_id)
            cursor.execute('''
                SELECT u.telegram_id FROM registrations r
                JOIN users u ON r.user_id = u.id
                WHERE r.olympiad_id = ? AND r.payment_status = 'approved'
            ''', (olympiad_id,))
        except ValueError:
            bot.send_message(user_id, "Xato: noto'g'ri olimpiada IDsi. 🤔")
            clear_user_state(user_id)
            return

    users = cursor.fetchall()
    conn.close()
    
    sent_count = 0
    error_count = 0
    
    for user_row in users:
        telegram_id = user_row[0]
        try:
            bot.send_message(telegram_id, text_to_send)
            sent_count += 1
            time.sleep(0.1)
        except Exception as e:
            logger.error(f"Failed to send broadcast to user {telegram_id}: {e}")
            error_count += 1
            
    bot.send_message(message.chat.id, f"Xabar yuborildi! ✅\nMuvaffaqiyatli: {sent_count} 🚀\nXatolik bilan: {error_count} 💔")
    
    clear_user_state(user_id)


# --- Start reminder system in a separate thread ---

reminder_thread = threading.Thread(target=send_reminders, daemon=True)
reminder_thread.start()

# --- Delete webhook before starting polling ---

try:
    bot.delete_webhook()
    print("Webhook deleted successfully")
except Exception as e:
    print(f"Error deleting webhook: {e}")

# --- Start the bot ---

if __name__ == "__main__":
    print("Bot is running...")
    try:
        bot.infinity_polling()
    except Exception as e:
        print(f"Bot error: {e}")
